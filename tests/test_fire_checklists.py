"""Fire checklists (PIL/EHS/CL 025-028) — offline tests (house no-DB style).

What is worth testing here is not that a checklist saves. It is that the four
things a controlled document promises actually hold:

  1. **The transcription is faithful.** Item counts, document numbers, revisions
     and effective dates match the client workbooks. A silently dropped item is
     the failure mode nobody notices until an auditor counts the rows.
  2. **A period resolves to exactly one record.** Period labels are canonical and
     malformed ones are rejected, because "2026-8" alongside "2026-08" is two
     Augusts and no error.
  3. **The sign-off chain cannot be skipped**, and an approved record cannot be
     edited.
  4. **The export renders**, in every layout, including the awkward cases — a
     31-column landscape month, an empty sheet, a shutdown week.

Everything here runs without a database.
"""

from __future__ import annotations

from datetime import date, datetime, timedelta, timezone

import pytest

from app.models.cams import CamsEngagement, CamsTemplate, CamsTemplateQuestion, CamsTemplateSection
from app.services import fire_checklist_pdf as pdfsvc
from app.services import fire_checklist_xlsx as xlsxsvc
from app.services import fire_checklists as svc
from app.services import fire_register as regsvc
from app.services.fire_checklist_templates import (
    ALL_TEMPLATES, BY_CODE, FE_REGISTER_DOC, FAS_MONTHLY_21A, FAS_MONTHLY_21B,
)

NOW = datetime(2026, 8, 19, tzinfo=timezone.utc)


# ═══════════════════════════════════════════════════════════════════════════
# 1. Transcription fidelity
# ═══════════════════════════════════════════════════════════════════════════
# (templateCode, item count, documentNo) read off the client workbooks.
EXPECTED = [
    ("PIL-FAS-DAILY", 7, "PIL/EHS/CL/025-R1 (A)"),
    ("PIL-FAS-MONTHLY-UNIT_21_A", 28, "PIL/EHS/CL/025-R1 (B)"),
    ("PIL-FAS-MONTHLY-UNIT_21_B", 27, "PIL/EHS/CL/025-R1 (B)"),
    ("PIL-FAS-QUARTERLY", 32, "PIL/EHS/CL/025-R1 (C)"),
    ("PIL-FAS-ANNUAL", 34, "PIL/EHS/CL/025-R1 (D)"),
    ("PIL-FAS-BEAM-DAILY", 5, "PIL/EHS/CL/025-R1 (E)"),
    ("PIL-FHS-DAILY", 8, "PIL/EHSD/CL/026-R2"),
    ("PIL-FHS-MONTHLY", 20, "PIL/EHSD/CL/026-R2"),
    ("PIL-FHS-QUARTERLY", 4, "PIL/EHSD/CL/026-R1"),
    ("PIL-FHS-YEARLY", 3, "PIL/EHSD/CL/026-R1 (D)"),
    ("PIL-FE-INSPECTION", 21, "PIL/EHSD/CL/027-R1"),
]


def test_all_eleven_sheets_are_defined():
    assert len(ALL_TEMPLATES) == 11
    assert {t.code for t in ALL_TEMPLATES} == {c for c, _n, _d in EXPECTED}


@pytest.mark.parametrize("code,count,doc_no", EXPECTED)
def test_item_count_and_document_number(code, count, doc_no):
    t = BY_CODE[code]
    assert len(t.items) == count, f"{code} item count drifted from the source sheet"
    assert t.documentNo == doc_no


def test_item_keys_are_unique_within_a_template():
    # Keys are the identity that survives a re-seed. A duplicate would silently
    # merge two rows of the sheet into one question and lose one of them.
    for t in ALL_TEMPLATES:
        keys = [i.key for i in t.items]
        assert len(set(keys)) == len(keys), f"{t.code} has duplicate item keys"


def test_fe_inspection_has_the_twenty_one_checks_in_sheet_order():
    t = BY_CODE["PIL-FE-INSPECTION"]
    keys = [i.key for i in t.items]
    assert keys == [f"fe_{n:02d}" for n in range(1, 22)]
    assert t.items[0].text.startswith("Exterior body")
    assert t.items[-1].text.startswith("Operation instructions sticker")


def test_every_pass_fail_check_raises_a_finding():
    """A "No" on a judgement check must lead somewhere. Readings must not.

    The exemptions are the point of this test: a battery voltage or a detector
    serial number has no "No" to fail, so raising a finding from one would be
    noise. Everything that is genuinely a pass/fail judgement raises.
    """
    # A Yes/No/NA check may be exempt, but only with a stated reason. An exemption
    # nobody had to justify is how the whole feature quietly reverts.
    unjustified = [
        (t.code, i.key)
        for t in ALL_TEMPLATES
        for i in t.items
        if not i.triggers_finding and i.type == "YES_NO_NA" and not i.no_finding_reason
    ]
    assert unjustified == [], f"pass/fail checks exempt from CAPA with no reason given: {unjustified}"

    # The only exemptions today are the inverted-polarity ones, where "No" is the
    # healthy answer and a CAPA on it would be backwards.
    justified = {
        i.key for t in ALL_TEMPLATES for i in t.items
        if not i.triggers_finding and i.type == "YES_NO_NA"
    }
    assert justified == {"fas_m_hooter_addl_required", "fas_q_hooter_addl_required"}

    raising = [i for t in ALL_TEMPLATES for i in t.items if i.triggers_finding]
    assert all(i.type == "YES_NO_NA" for i in raising)
    assert len(raising) > 100, "the bulk of the checks should raise"


def test_severity_is_reserved_for_checks_that_mean_the_system_is_dead():
    """MAJOR_NC has to stay scarce or it stops meaning anything."""
    majors = {i.key for t in ALL_TEMPLATES for i in t.items if i.nc_severity == "MAJOR_NC"}
    # The ones a fire officer would stop the line for.
    for key in ("fas_d_03", "fas_d_07", "fhs_d_02", "fhs_d_07", "fe_06", "fe_14"):
        assert key in majors, f"{key} should be MAJOR_NC — a 'No' means it will not work in a fire"
    # And the ones that are housekeeping must NOT be.
    minors = {i.key: i.nc_severity for t in ALL_TEMPLATES for i in t.items}
    assert minors["fe_02"] == "MINOR_NC", "paint condition is not a major non-conformance"
    assert minors["fe_17"] == "MINOR_NC", "a missing signage board is not a major non-conformance"
    total = sum(1 for t in ALL_TEMPLATES for i in t.items if i.triggers_finding)
    assert len(majors) < total * 0.25, (
        f"{len(majors)} of {total} raising checks are MAJOR_NC — severity is losing its meaning"
    )


def test_the_two_unit_monthly_sheets_differ_only_in_addressing_and_hooters():
    """Unit-21 A is a ZONE panel, Unit-21 B is a LOOP panel.

    This is the whole reason `siteVariant` exists, so it is worth asserting that
    the difference is real and is confined to the two places it should be — if a
    future edit made the sheets identical, one unit would be inspected against
    the wrong document and nothing else would complain.
    """
    a = {i.key: i.text for i in FAS_MONTHLY_21A.items}
    b = {i.key: i.text for i in FAS_MONTHLY_21B.items}
    assert a["fas_m_trig_addr"].startswith("Zone Number")
    assert b["fas_m_trig_addr"].startswith("Loop Number")
    assert FAS_MONTHLY_21A.siteVariant == "UNIT_21_A"
    assert FAS_MONTHLY_21B.siteVariant == "UNIT_21_B"
    # Same document number, different content — that is the point of the variant.
    assert FAS_MONTHLY_21A.documentNo == FAS_MONTHLY_21B.documentNo

    hooters_a = {k for k in a if k.startswith("fas_m_hooter_") and k[-2:].isdigit()}
    hooters_b = {k for k in b if k.startswith("fas_m_hooter_") and k[-2:].isdigit()}
    assert len(hooters_a) == 6 and len(hooters_b) == 5

    # Everything that is NOT the addressing field or a hooter must be identical.
    shared = (set(a) & set(b)) - {"fas_m_trig_addr"} - hooters_a - hooters_b
    assert all(a[k] == b[k] for k in shared)


def test_register_document_has_the_sixteen_sheet_columns():
    assert len(FE_REGISTER_DOC["columns"]) == 16
    assert FE_REGISTER_DOC["documentNo"] == "PIL/EHSD/CL/028-R1"
    labels = [lab for _k, lab in FE_REGISTER_DOC["columns"]]
    assert labels[:4] == ["Sl. No", "Manufacturer Serial No.", "Type", "Capacity"]
    assert labels[-1] == "Remarks"


# ═══════════════════════════════════════════════════════════════════════════
# 2. Period identity
# ═══════════════════════════════════════════════════════════════════════════
@pytest.mark.parametrize(
    "freq,expected",
    [("DAILY", "2026-08-19"), ("MONTHLY", "2026-08"), ("QUARTERLY", "2026-Q3"), ("ANNUAL", "2026")],
)
def test_period_label_is_canonical(freq, expected):
    assert svc.period_label(freq, date(2026, 8, 19)) == expected


@pytest.mark.parametrize(
    "freq,label",
    [
        ("MONTHLY", "2026-8"),      # unpadded — would be a second August
        ("DAILY", "2026-13-01"),    # impossible month
        ("DAILY", "2026-02-30"),    # not a real date
        ("QUARTERLY", "2026-Q5"),
        ("ANNUAL", "26"),
        ("DAILY", ""),
    ],
)
def test_malformed_periods_are_rejected(freq, label):
    with pytest.raises(svc.ChecklistError):
        svc.validate_period(freq, label)


def test_quarter_period_starts_on_the_quarter():
    assert svc.period_start("QUARTERLY", "2026-Q3") == datetime(2026, 7, 1, tzinfo=timezone.utc)
    assert svc.period_start("ANNUAL", "2026") == datetime(2026, 1, 1, tzinfo=timezone.utc)


def test_day_grid_columns_track_the_real_month_length():
    assert len(svc.grid_periods("DAY_GRID", "DAILY", "2026-02")) == 28
    assert len(svc.grid_periods("DAY_GRID", "DAILY", "2024-02")) == 29   # leap
    assert len(svc.grid_periods("DAY_GRID", "DAILY", "2026-08")) == 31


def test_month_and_quarter_grids_are_paged_by_year():
    months = svc.grid_periods("MONTH_GRID", "MONTHLY", "2026")
    assert [h for _p, h in months] == ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                                       "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    quarters = svc.grid_periods("QUARTER_GRID", "QUARTERLY", "2026")
    assert [p for p, _h in quarters] == ["2026-Q1", "2026-Q2", "2026-Q3", "2026-Q4"]
    # The sheet's own column captions, not "Q1".
    assert [h for _p, h in quarters][0] == "First Quarter"


def test_grid_window_must_match_the_layout():
    with pytest.raises(svc.ChecklistError):
        svc.grid_periods("DAY_GRID", "DAILY", "2026")      # a daily grid pages by month
    with pytest.raises(svc.ChecklistError):
        svc.grid_periods("MONTH_GRID", "MONTHLY", "2026-08")  # a month grid pages by year


def test_window_paging_wraps_the_year_boundary():
    assert svc.shift_window("DAY_GRID", "2026-01", -1) == "2025-12"
    assert svc.shift_window("DAY_GRID", "2026-12", 1) == "2027-01"
    assert svc.shift_window("MONTH_GRID", "2026", 1) == "2027"


# ═══════════════════════════════════════════════════════════════════════════
# 3. Sign-off chain and immutability
# ═══════════════════════════════════════════════════════════════════════════
def _run(status: str) -> CamsEngagement:
    e = CamsEngagement()
    e.status = status
    return e


@pytest.mark.parametrize(
    "status,stage",
    [
        ("PLANNED", "DRAFT"), ("SCHEDULED", "DRAFT"), ("IN_PROGRESS", "DRAFT"),
        ("FIELDWORK_COMPLETE", "SUBMITTED"), ("FINDINGS_REVIEW", "REVIEWED"),
        ("REPORT_ISSUED", "APPROVED"), ("CLOSED", "APPROVED"),
    ],
)
def test_cams_status_maps_to_a_sign_off_stage(status, stage):
    assert svc.stage_of(_run(status)) == stage


def test_only_a_draft_is_editable():
    assert not svc.is_locked(_run("IN_PROGRESS"))
    for status in ("FIELDWORK_COMPLETE", "FINDINGS_REVIEW", "REPORT_ISSUED", "CLOSED"):
        assert svc.is_locked(_run(status)), f"{status} must be locked"


def test_each_stage_requires_its_exact_predecessor():
    """No skipping. Approve-before-review must be impossible, not merely unlikely."""
    assert svc._STAGE_STEP[svc.STAGE_SUBMITTED][0] == svc.STAGE_DRAFT
    assert svc._STAGE_STEP[svc.STAGE_REVIEWED][0] == svc.STAGE_SUBMITTED
    assert svc._STAGE_STEP[svc.STAGE_APPROVED][0] == svc.STAGE_REVIEWED
    # DRAFT is not a destination — there is no un-submit.
    assert svc.STAGE_DRAFT not in svc._STAGE_STEP


def test_run_payload_reports_whether_a_signature_is_required():
    """`signatureRequired` must carry the template's real answer, not None.

    It shipped as a hardcoded None and every caller reads it as `!== false`, so a
    DAILY sheet — which `signature_enforced` deliberately exempts, because 31 drawn
    signatures a month gets the tablet handed round — still demanded one on screen.
    """
    monthly = _template(("a", "Panel normal", True))
    assert svc.signature_enforced(monthly) is True

    daily = _template(("a", "Panel normal", True))
    daily.documentMeta = {**daily.documentMeta, "frequency": "DAILY", "layout": "DAY_GRID"}
    assert svc.signature_enforced(daily) is False

    run = _run("IN_PROGRESS")
    run.id = "r1"
    run.periodLabel = "2026-08"
    run.siteId = "p1"
    run.engagementCode = "INS-1"
    run.sourceEntityId = None
    run.areaOrAssetRef = "FE-1"
    run.reviewedBy = run.approvedBy = run.reviewedAt = run.approvedAt = None
    run.scorePercent = run.overallResult = None

    for tpl, expected in ((monthly, True), (daily, False)):
        payload = svc.run_out(tpl, run, None, None)
        assert payload["signOff"]["signatureRequired"] is expected


def test_approval_stops_at_report_issued_not_closed():
    """Approving the sheet must not close out the defects it raised.

    CLOSED runs the close-blocker check on open findings. Collapsing approve and
    close would let an HOD sign a defect away by approving the sheet it was found
    on, which is the opposite of what the sign-off is for.
    """
    assert svc._STAGE_STEP[svc.STAGE_APPROVED][1] == "REPORT_ISSUED"


def _template(*items: tuple[str, str, bool]) -> CamsTemplate:
    """Detached template: (key, text, mandatory)."""
    tpl = CamsTemplate()
    tpl.templateCode = "T"
    tpl.documentMeta = {"documentNo": "X", "frequency": "MONTHLY", "layout": "FORM"}
    sec = CamsTemplateSection()
    sec.title = "Checks"
    sec.orderIndex = 0
    sec.questions = []
    for i, (key, text, mandatory) in enumerate(items):
        q = CamsTemplateQuestion()
        q.id = f"q{i}"
        q.standardClauseRef = key
        q.text = text
        q.orderIndex = i
        q.questionType = "YES_NO_NA"
        q.isMandatory = mandatory
        sec.questions.append(q)
    tpl.sections = [sec]
    return tpl


class _Resp:
    def __init__(self, answers):
        self.answers = answers


def test_submit_gate_lists_unanswered_mandatory_items():
    tpl = _template(("a", "Panel normal", True), ("b", "Optional note", False), ("c", "Hooter works", True))
    missing = svc.unanswered_mandatory(tpl, _Resp([{"questionId": "q0", "value": "YES"}]))
    assert missing == ["Hooter works"]


def test_blank_is_not_the_same_as_na():
    """An inspector who has not looked and one who found the check inapplicable
    are recording different facts. A sheet that conflates them cannot be audited."""
    tpl = _template(("a", "Panel normal", True))
    assert svc.unanswered_mandatory(tpl, _Resp([{"questionId": "q0", "value": None}])) == ["Panel normal"]
    assert svc.unanswered_mandatory(tpl, _Resp([{"questionId": "q0", "value": ""}])) == ["Panel normal"]
    assert svc.unanswered_mandatory(tpl, _Resp([{"questionId": "q0", "value": "NA"}])) == []


def test_sheet_vocabulary_maps_onto_engine_conformance():
    # The sheets say Yes/No/NA; CAMS scores CONFORM/NC/NA. Same three states.
    assert svc._VALUE_TO_CONFORMANCE == {"YES": "CONFORM", "NO": "NC", "NA": "NA"}
    assert svc.CONFORMANCE_TO_VALUE["NC"] == "NO"


def _question(qtype: str) -> CamsTemplateQuestion:
    q = CamsTemplateQuestion()
    q.questionType = qtype
    q.text = "Q"
    return q


def test_yes_no_na_answers_are_validated():
    q = _question("YES_NO_NA")
    assert svc._coerce(q, "yes") == ("YES", "CONFORM")
    assert svc._coerce(q, "NO") == ("NO", "NC")
    assert svc._coerce(q, None) == (None, None)
    with pytest.raises(svc.ChecklistError):
        svc._coerce(q, "MAYBE")


def test_readings_carry_no_conformance():
    """A battery voltage is a reading, not a judgement.

    Scoring skips answers with no conformance, so a monthly sheet's score
    reflects its pass/fail checks and is not diluted by someone writing down 27.4.
    """
    assert svc._coerce(_question("NUMERIC"), "27.4") == ("27.4", None)
    assert svc._coerce(_question("TEXT"), " SD-14 ") == ("SD-14", None)
    with pytest.raises(svc.ChecklistError):
        svc._coerce(_question("NUMERIC"), "twenty seven")


# ═══════════════════════════════════════════════════════════════════════════
# 4. Register badges
# ═══════════════════════════════════════════════════════════════════════════
def test_badge_ladder():
    assert regsvc.badge_for(NOW - timedelta(days=1), NOW)["status"] == "OVERDUE"
    assert regsvc.badge_for(NOW + timedelta(days=10), NOW)["status"] == "DUE_SOON"
    assert regsvc.badge_for(NOW + timedelta(days=30), NOW)["status"] == "DUE_SOON"
    assert regsvc.badge_for(NOW + timedelta(days=31), NOW)["status"] == "OK"


def test_a_missing_due_date_is_a_register_gap_not_compliance():
    """The one failure mode that makes a due-date register worse than none."""
    b = regsvc.badge_for(None, NOW)
    assert b["status"] == "NOT_RECORDED"
    assert b["status"] != "OK"


def test_badge_reports_days_remaining_signed():
    assert regsvc.badge_for(NOW - timedelta(days=5), NOW)["daysRemaining"] == -5
    assert regsvc.badge_for(NOW + timedelta(days=5), NOW)["daysRemaining"] == 5


# ═══════════════════════════════════════════════════════════════════════════
# 5. PDF export
# ═══════════════════════════════════════════════════════════════════════════
DOC = {
    "documentNo": "PIL/EHS/CL/025-R1 (A)", "supersedesNo": "PIL/EHS/CL/002-R0", "revision": "R1",
    "effectiveDate": "2025-04-01", "reviewDate": "2028-03-31", "department": "EHS",
    "signOffRoles": ["Prepared by: Person In-charge", "Reviewed by: Intermediatory Head",
                     "Approved by: HOD"],
    "footnotes": ["Electrician Signature", "Safety Officer signature"],
}


def _grid_payload(n_cols: int, *, shutdown: bool = False, empty: bool = False) -> dict:
    cols = [
        {
            "periodLabel": f"2026-08-{d:02d}", "header": str(d), "runId": None,
            "stage": None if empty else ("APPROVED" if d <= 3 else "DRAFT"),
            "locked": not empty and d <= 3,
            "nonWorkingDay": "SUNDAY" if (shutdown and d % 7 == 2) else None,
        }
        for d in range(1, n_cols + 1)
    ]
    rows = [
        {
            "questionId": f"q{i}", "itemKey": f"fas_d_0{i}", "sectionTitle": "Daily Attention",
            "text": "Check the pump glands, packing's, etc., and replace the damaged gland for "
                    "packing whenever found damaged or worn out.",
            "type": "YES_NO_NA", "guidance": "Leakage from glands at OEM rate is allowable",
            "cells": {
                c["periodLabel"]: {"value": None if empty else ("NO" if i == 3 else "YES"),
                                   "conformance": None, "note": ""}
                for c in cols
            },
        }
        for i in range(1, 8)
    ]
    return {
        "document": DOC, "templateName": "Daily Fire Alarm System Inspection Checklist",
        "layout": "DAY_GRID", "window": "2026-08", "assetCode": "FIRE-P1-FAS-A",
        "assetLocation": "Unit-21 A - Panel Room", "assetSubtype": "ZONE",
        "allottedSerialNo": None, "columns": cols, "rows": rows,
    }


@pytest.mark.parametrize(
    "kwargs",
    [{}, {"shutdown": True}, {"empty": True}],
    ids=["filled", "shutdown-week", "never-inspected"],
)
def test_day_grid_renders(kwargs):
    pdf = pdfsvc.render_grid(_grid_payload(31, **kwargs))
    assert pdf.startswith(b"%PDF") and len(pdf) > 1500


def test_quarter_grid_renders_portrait():
    payload = _grid_payload(4)
    payload["layout"] = "QUARTER_GRID"
    payload["columns"] = [
        {**c, "header": h, "periodLabel": f"2026-Q{i}"}
        for i, (c, h) in enumerate(
            zip(payload["columns"], ["First Quarter", "Second Quarter", "Third Quarter", "Forth Quarter"]),
            start=1,
        )
    ]
    for r in payload["rows"]:
        r["cells"] = {c["periodLabel"]: {"value": "YES", "conformance": None, "note": ""}
                      for c in payload["columns"]}
    assert pdfsvc.render_grid(payload).startswith(b"%PDF")


def test_form_renders_with_a_partial_sign_off():
    """An unsigned stage must print blank, not a name we could have guessed."""
    payload = {
        "document": {**DOC, "documentNo": "PIL/EHS/CL/025-R1 (B)"},
        "templateName": "AFDAS Monthly (Unit-21 A)", "assetCode": "FIRE-P1-FAS-A",
        "assetLocation": "Unit-21 A", "periodLabel": "2026-08", "stage": "SUBMITTED",
        "signOff": {
            "preparedByName": "R Kumar", "preparedAt": "2026-08-31T05:00:00+00:00",
            "reviewedByName": None, "reviewedAt": None,
            "approvedByName": None, "approvedAt": None, "roles": DOC["signOffRoles"],
        },
        "sections": [
            {"title": "Monthly Attention", "note": None, "items": [
                {"text": "Manual call point is in working condition.", "type": "YES_NO_NA",
                 "guidance": None, "value": "NO", "note": "MCP-4 lens cracked"},
            ]},
            {"title": "Battery Details", "note": "20 % of ___ Detectors", "items": [
                {"text": "Battery (B1) = ______ VDC", "type": "NUMERIC", "guidance": None,
                 "value": "27.4", "note": ""},
            ]},
        ],
    }
    assert pdfsvc.render_form(payload).startswith(b"%PDF")


def test_register_renders_every_badge_state():
    def row(n, life, hp, refill):
        return {
            "slNo": n, "serialNo": f"MFR-{n}", "type": "CO2", "capacity": "2KG",
            "yearOfManufacture": 2021, "expiryDate": "2031-04-27T00:00:00+00:00",
            "make": "SAFETECH", "allottedSerialNo": str(36770 + n),
            "location": "Admin - Reception, near the fire exit on the north side",
            "hpTestedOn": "2021-04-27T00:00:00+00:00", "hpTestDueDate": "2026-07-30T00:00:00+00:00",
            "dateOfDischarge": None, "refilledOn": None, "dueForRefilling": None,
            "weightKg": 2.0, "remarks": None,
            "badges": {"cylinderLife": {"status": life}, "hpTest": {"status": hp},
                       "refill": {"status": refill}},
            "worstBadge": hp,
        }

    payload = {
        "document": FE_REGISTER_DOC,
        "summary": {"total": 4, "overdue": 1, "dueSoon": 1, "notRecorded": 1},
        "rows": [
            row(1, "OK", "OVERDUE", "OK"),
            row(2, "OK", "DUE_SOON", "NOT_RECORDED"),
            row(3, "OK", "OK", "OK"),
            row(4, "OVERDUE", "NOT_RECORDED", "NOT_RECORDED"),
        ],
    }
    assert pdfsvc.render_register(payload).startswith(b"%PDF")


def test_register_renders_when_empty():
    payload = {"document": FE_REGISTER_DOC,
               "summary": {"total": 0, "overdue": 0, "dueSoon": 0, "notRecorded": 0}, "rows": []}
    assert pdfsvc.render_register(payload).startswith(b"%PDF")


# ═══════════════════════════════════════════════════════════════════════════
# 5. The Excel export
# ═══════════════════════════════════════════════════════════════════════════
# The workbook is the working copy, not a second controlled document, so what is
# worth asserting is different from the PDF's "it rendered": that it OPENS, that
# it keeps one row per check and one column per period, and — the thing that
# would regress silently — that a due date arrives as a real date Excel can sort
# and a missing one still says "not recorded" rather than going blank.
def _load(data: bytes):
    from io import BytesIO

    from openpyxl import load_workbook

    assert data.startswith(b"PK"), "an .xlsx is a zip; anything else Excel will refuse"
    return load_workbook(BytesIO(data)).active


@pytest.mark.parametrize(
    "kwargs",
    [{}, {"shutdown": True}, {"empty": True}],
    ids=["filled", "shutdown-week", "never-inspected"],
)
def test_day_grid_xlsx_opens_with_a_column_per_day(kwargs):
    payload = _grid_payload(31, **kwargs)
    ws = _load(xlsxsvc.render_grid(payload))
    # Sl. + wording + 31 days. A column silently dropped from a month page is the
    # failure an inspector only finds when the 31st has nowhere to go.
    assert ws.max_column == 2 + 31
    assert ws.freeze_panes  # the wording must stay put when scrolling to the 31st


def test_grid_exports_carry_the_remark_that_explains_a_no():
    """A "No" in a grid cell is a question; the remark is the answer.

    The cell is eight millimetres wide and the source sheet's own footnote sends
    the inspector to "the back side of this page" for exactly this reason. Both
    exports have to reproduce that back page, or a NO prints as the bare word and
    whoever reads it knows an item failed but not what was seen.
    """
    remark = "Number plate painted over during last repaint."
    payload = _grid_payload(31)
    first = payload["rows"][0]
    period = payload["columns"][0]["periodLabel"]
    first["cells"][period] = {"value": "NO", "conformance": "NC", "note": remark}

    ws = _load(xlsxsvc.render_grid(payload))
    flat = [str(v) for row in ws.iter_rows(values_only=True) for v in row if v is not None]
    assert any(remark in v for v in flat), "the Remarks block must list it"
    comments = [c.comment.text for r in ws.iter_rows() for c in r if c.comment]
    assert any(remark in t for t in comments), "and the cell itself must carry it"

    # The PDF grows: the remarks block is rendered, not silently dropped.
    with_remark = pdfsvc.render_grid(payload)
    first["cells"][period] = {"value": "NO", "conformance": "NC", "note": ""}
    without = pdfsvc.render_grid(payload)
    assert with_remark.startswith(b"%PDF") and len(with_remark) > len(without)


def test_grid_export_has_no_remarks_block_when_nothing_was_written():
    """An empty Remarks heading on a clean month reads as a missing comment."""
    ws = _load(xlsxsvc.render_grid(_grid_payload(31)))
    flat = [str(v) for row in ws.iter_rows(values_only=True) for v in row if v is not None]
    assert not any(v == "Remarks" for v in flat)


def test_form_xlsx_keeps_every_item_and_its_note():
    payload = {
        "document": {**DOC, "documentNo": "PIL/EHS/CL/025-R1 (B)"},
        "templateName": "AFDAS Monthly (Unit-21 A)", "assetCode": "FIRE-P1-FAS-A",
        "assetLocation": "Unit-21 A", "periodLabel": "2026-08", "stage": "SUBMITTED",
        "signOff": {
            "preparedByName": "R Kumar", "preparedAt": "2026-08-31T05:00:00+00:00",
            "reviewedByName": None, "reviewedAt": None,
            "approvedByName": None, "approvedAt": None, "roles": DOC["signOffRoles"],
        },
        "sections": [
            {"title": "Monthly Attention", "note": None, "items": [
                {"text": "Manual call point is in working condition.", "type": "YES_NO_NA",
                 "guidance": None, "value": "NO", "note": "MCP-4 lens cracked"},
            ]},
        ],
    }
    ws = _load(xlsxsvc.render_form(payload))
    flat = [str(v) for row in ws.iter_rows(values_only=True) for v in row if v is not None]
    assert any("Manual call point" in v for v in flat)
    # The observation and the inspector's note both have to survive — a "NO" with
    # its reason stripped is the half of the record that matters.
    assert "NO" in flat and any("MCP-4 lens cracked" in v for v in flat)


def test_register_xlsx_writes_real_dates_and_keeps_not_recorded():
    from datetime import date as _date

    def row(n, life, hp, refill, refill_on=None, refill_due=None):
        return {
            "slNo": n, "serialNo": f"MFR-{n}", "type": "CO2", "capacity": "2KG",
            "yearOfManufacture": 2021, "expiryDate": "2031-04-27T00:00:00+00:00",
            "make": "SAFETECH", "allottedSerialNo": str(36770 + n),
            "location": "Admin - Reception", "hpTestedOn": "2021-04-27T00:00:00+00:00",
            "hpTestDueDate": "2026-07-30T00:00:00+00:00", "dateOfDischarge": None,
            "refilledOn": refill_on, "dueForRefilling": refill_due,
            "weightKg": 2.0, "remarks": None,
            "badges": {"cylinderLife": {"status": life}, "hpTest": {"status": hp},
                       "refill": {"status": refill}},
            "worstBadge": hp,
        }

    payload = {
        "document": FE_REGISTER_DOC,
        "summary": {"total": 2, "overdue": 1, "dueSoon": 0, "notRecorded": 1},
        "rows": [
            row(1, "OK", "OVERDUE", "OK", "2025-04-11T00:00:00+00:00", "2028-04-10T00:00:00+00:00"),
            row(2, "OK", "OVERDUE", "NOT_RECORDED"),
        ],
    }
    ws = _load(xlsxsvc.render_register(payload))
    values = [c.value for r in ws.iter_rows() for c in r]
    # A real date, not "10.04.2028" as text — the register is read by due date, and
    # a string column sorts 01.02.2031 above 27.04.2026. openpyxl reads a date cell
    # back as a datetime, so compare on the date part.
    dates = {v.date() if isinstance(v, datetime) else v for v in values if isinstance(v, (_date, datetime))}
    assert _date(2028, 4, 10) in dates
    # A cylinder with no refill date on file is a register GAP. Blank would read
    # as "nothing due", which is the one thing this column must never say.
    assert "not recorded" in values
    assert ws.auto_filter.ref, "the filter is the reason this is a workbook and not a second PDF"


def test_register_xlsx_opens_when_empty():
    payload = {"document": FE_REGISTER_DOC,
               "summary": {"total": 0, "overdue": 0, "dueSoon": 0, "notRecorded": 0}, "rows": []}
    assert _load(xlsxsvc.render_register(payload)) is not None


def test_asset_register_xlsx_and_pdf_render():
    """The 'All other fire assets' tab — panels, hydrants, detectors."""
    rows = [
        {"equipmentCode": "FE-AGB-0006", "type": "FIRE_ALARM_PANEL", "assetSubtype": None,
         "location": "Admin - Lobby", "capacitySpec": None, "make": None, "model": None,
         "serialNo": None, "maintenanceContractor": "SafeFire Services Pvt Ltd",
         "lastInspectionDate": "2026-06-15T19:08:30+00:00",
         "nextInspectionDueDate": "2026-07-15T19:08:30+00:00", "status": "OVERDUE"},
        {"equipmentCode": "FIRE-ACS-FHS-01", "type": "FIRE_HYDRANT_SYSTEM", "assetSubtype": None,
         "location": "Fire Pump House - Main Yard", "capacitySpec": "Hydrant & Sprinkler System",
         "make": None, "model": None, "serialNo": None, "maintenanceContractor": None,
         "lastInspectionDate": None, "nextInspectionDueDate": None, "status": "DUE_INSPECTION"},
    ]
    ws = _load(xlsxsvc.render_assets(rows))
    flat = [c.value for r in ws.iter_rows() for c in r]
    assert "FE-AGB-0006" in flat and "FIRE-ACS-FHS-01" in flat
    # An asset that has never been inspected must still render — it is precisely
    # the row someone is looking for.
    assert pdfsvc.render_assets(rows).startswith(b"%PDF")


def test_asset_register_renders_when_empty():
    assert _load(xlsxsvc.render_assets([])) is not None
    assert pdfsvc.render_assets([]).startswith(b"%PDF")
