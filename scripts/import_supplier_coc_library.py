"""Import the Supplier Code of Conduct checklist as a VENDOR-scoped library.

Source: `docs/sheet/supplier-compliance-checklist-starter.xlsx`, sheet
"Supplier Compliance Checklist" — 39 checkpoints across 5 pillars, SA8000 /
SMETA-aligned.

**This is demo starter content and is labelled as such in the data**, not just
in a document. Every checkpoint carries
`citation_status = UNVERIFIED_STARTER_CONTENT`, so the report footnote counts it
among the unverified citations and nobody downstream mistakes "SA8000 §5" here
for a sourced, verified reference. The workbook's own first row says "replace
with the client's own Supplier Code of Conduct checklist when available"; that
instruction is worth nothing if it stops at the spreadsheet.

**Why one library rather than one per industry.** The `Applicable Industries`
column is preserved as per-checkpoint metadata (`applicable_industries`), not
used to split the content. Splitting would produce five near-duplicate libraries
that drift apart the moment anyone edits one.

Mirrors `seed-audit-compliance.ts` exactly for structure, the `cp()` rule-fill
(critical/major fails need a photo; criticals auto-spawn a CAPA) and count
computation — `checkpointCount` is derived by `import_library`, the same path the
industry libraries use.

Idempotent (upsert by industryCode). Dry run by default.

    .venv/Scripts/python.exe scripts/import_supplier_coc_library.py
    .venv/Scripts/python.exe scripts/import_supplier_coc_library.py --commit

WARNING: the backend .env points at PRODUCTION.
"""

from __future__ import annotations

import argparse
import sys
from collections import Counter
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.models.audit_compliance import AuditCheckpointLibrary
from app.services import citations as cit
from app.services.audit_compliance import library_subject_scope

XLSX = (
    Path(__file__).resolve().parents[2]
    / "docs" / "sheet" / "supplier-compliance-checklist-starter.xlsx"
)
SHEET = "Supplier Compliance Checklist"

INDUSTRY_CODE = "SUPPLIER_COC"
INDUSTRY_NAME = "Supplier Code of Conduct"
VERSION = "2026.1-starter"

SOURCE_TAG = (
    "SafeOps360 starter library (SA8000/SMETA-aligned), imported from "
    "supplier-compliance-checklist-starter.xlsx — DEMO CONTENT, replace with the "
    "client's own Supplier Code of Conduct when available."
)

# Pillar -> the presentation the audit UI already expects on a category. Colours
# follow the existing libraries' convention (one distinct hue per discipline).
PILLARS: dict[str, dict[str, str]] = {
    "Labour Standards":  {"code": "SUP-LABOUR",   "color": "#7C3AED", "icon": "users"},
    "Health & Safety":   {"code": "SUP-HS",       "color": "#DC2626", "icon": "shield"},
    "Environment":       {"code": "SUP-ENV",      "color": "#059669", "icon": "leaf"},
    "Business Ethics":   {"code": "SUP-ETHICS",   "color": "#0891B2", "icon": "scale"},
    "Management System": {"code": "SUP-MGMT",     "color": "#D97706", "icon": "clipboard"},
}

# The workbook's Title-case severities -> the audit engine's EXISTING lowercase
# checkpoint vocabulary (`AuditCheckpointResponse.criticality`, default "major";
# `_CAPA_SEVERITY` maps the same three). Deliberately NOT the inspection
# engine's `CRITICAL_NC/MAJOR_NC/...` — that is a different model's vocabulary,
# and introducing a third would be the drift this mapping exists to prevent.
SEVERITY = {"critical": "critical", "major": "major", "minor": "minor"}


def read_rows() -> list[dict[str, str]]:
    if not XLSX.exists():
        print(f"ERROR: workbook not found at {XLSX}", file=sys.stderr)
        raise SystemExit(2)
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    if SHEET not in wb.sheetnames:
        print(f"ERROR: sheet '{SHEET}' not in {wb.sheetnames}", file=sys.stderr)
        raise SystemExit(2)
    # The 'Notes' sheet is human documentation and is deliberately never read.
    rows = list(wb[SHEET].iter_rows(values_only=True))
    # Two banner rows precede the header, so the header is found rather than
    # assumed — an added banner row would otherwise shift every column silently.
    try:
        h = next(i for i, r in enumerate(rows) if (r[0] or "") == "ID")
    except StopIteration:
        print("ERROR: no header row with an 'ID' column", file=sys.stderr)
        raise SystemExit(2)
    keys = [str(c).strip() if c else "" for c in rows[h]]
    out = []
    for r in rows[h + 1:]:
        if not r[0]:
            continue
        out.append({k: (str(v).strip() if v is not None else "") for k, v in zip(keys, r)})
    return out


def build_payload(rows: list[dict[str, str]]) -> dict[str, Any]:
    by_pillar: dict[str, list[dict[str, Any]]] = {p: [] for p in PILLARS}
    unknown_pillars: set[str] = set()
    bad_severity: list[str] = []

    for r in rows:
        pillar = r.get("Pillar", "")
        if pillar not in PILLARS:
            unknown_pillars.add(pillar)
            continue

        raw_sev = (r.get("Severity") or "").strip().lower()
        crit = SEVERITY.get(raw_sev)
        if crit is None:
            # Never guess a severity. An unmapped value would otherwise land on
            # the "major" default and quietly downgrade a critical finding.
            bad_severity.append(f"{r.get('ID')}={r.get('Severity')!r}")
            continue

        is_crit, is_major = crit == "critical", crit == "major"
        reference = r.get("Reference", "")
        by_pillar[pillar].append({
            "code": r["ID"],
            "question": r.get("Checkpoint", ""),
            "guidance": r.get("Expected Evidence", ""),
            "requirement_reference": reference,
            # Derived from the citation itself rather than hardcoded, so an
            # SA8000-referenced checkpoint reports SA8000 and a SMETA-only one
            # reports SMETA. The clause index groups on (standard, reference).
            "standard": "SA8000" if "SA8000" in reference else "SMETA",
            "criticality": crit,
            # Same rule-fill as `cp()` in seed-audit-compliance.ts.
            "response_type": "pass_partial_fail",
            "requires_photo_on_fail": is_crit or is_major,
            "auto_trigger_capa_on_fail": is_crit,
            "capa_severity_if_triggered": "critical" if is_crit else ("major" if is_major else "minor"),
            "linked_safeops_module": None,
            # Preserved as metadata only — no UI is built for it here.
            "applicable_industries": r.get("Applicable Industries", ""),
            # Provenance: a person wrote these, but nobody has checked them
            # against SA8000/SMETA, so they count as unverified in the report.
            cit.KEY_STATUS: cit.UNVERIFIED_STARTER_CONTENT,
            cit.KEY_SOURCE: SOURCE_TAG,
            cit.KEY_PRIORITY: cit.NORMAL,
        })

    if unknown_pillars:
        print(f"ERROR: unmapped pillar(s): {sorted(unknown_pillars)}", file=sys.stderr)
        raise SystemExit(2)
    if bad_severity:
        print(f"ERROR: unmapped severity value(s): {bad_severity}", file=sys.stderr)
        raise SystemExit(2)

    categories = []
    for i, (pillar, meta) in enumerate(PILLARS.items(), start=1):
        cps = by_pillar[pillar]
        if not cps:
            continue
        categories.append({
            "category_code": meta["code"],
            "category_name": pillar,
            "category_color": meta["color"],
            "category_icon": meta["icon"],
            "sequence": i,
            # THE tag that puts this library in the Supplier branch of the
            # scheduling wizard. Read by `library_subject_scope`; without it the
            # library would classify as own-facility and never be offered for a
            # supplier audit.
            "subject_scope": "VENDOR",
            "checkpoints": cps,
        })

    return {
        "industryCode": INDUSTRY_CODE,
        "industryName": INDUSTRY_NAME,
        "version": VERSION,
        "categories": categories,
    }


def main(commit: bool) -> int:
    rows = read_rows()
    print(f"-- source: {XLSX.name} · sheet '{SHEET}'")
    print(f"   {len(rows)} checkpoint row(s)")
    print(f"   pillars : {dict(Counter(r['Pillar'] for r in rows))}")
    print(f"   severity: {dict(Counter(r['Severity'] for r in rows))}")
    print()

    payload = build_payload(rows)
    total = sum(len(c["checkpoints"]) for c in payload["categories"])

    print("-- library to import ------------------------------------------")
    print(f"   industryCode : {payload['industryCode']}")
    print(f"   industryName : {payload['industryName']}")
    print(f"   subjectScope : {library_subject_scope(payload['industryCode'], payload['categories'])}")
    print(f"   checkpoints  : {total}")
    for c in payload["categories"]:
        counts = Counter(cp["criticality"] for cp in c["checkpoints"])
        print(f"     {c['category_name']:<20} {len(c['checkpoints']):>3}  "
              f"(critical {counts['critical']}, major {counts['major']}, minor {counts['minor']})")
    print()

    if total != len(rows):
        print(f"ERROR: {len(rows)} source rows but {total} built — refusing to import",
              file=sys.stderr)
        return 2

    engine = create_engine(get_settings().sync_database_url, future=True)
    with Session(engine) as s:
        # Snapshot the OTHER libraries so the import can prove it left them alone.
        before = {
            l.industryCode: l.checkpointCount
            for l in s.execute(select(AuditCheckpointLibrary)).scalars().all()
        }

        lib = s.execute(
            select(AuditCheckpointLibrary).where(
                AuditCheckpointLibrary.industryCode == INDUSTRY_CODE
            )
        ).scalar_one_or_none()

        if not commit:
            print("  DRY RUN — nothing written (pass --commit to apply)")
            print(f"  would {'REPLACE' if lib else 'CREATE'} {INDUSTRY_CODE}")
            return 0

        if lib is None:
            lib = AuditCheckpointLibrary(
                industryCode=INDUSTRY_CODE, industryName=INDUSTRY_NAME,
                version=VERSION, categories=payload["categories"],
                checkpointCount=total, isActive=True,
            )
            s.add(lib)
            action = "CREATED"
        else:
            lib.industryName = INDUSTRY_NAME
            lib.version = VERSION
            lib.categories = payload["categories"]
            lib.checkpointCount = total
            lib.isActive = True
            action = "REPLACED"
        s.commit()
        print(f"  {action} {INDUSTRY_CODE}")

    # ── Prove it, from a fresh session ────────────────────────────────
    # Same lesson as the clause import: an in-memory read cannot distinguish
    # "written" from "believed to be written".
    with Session(engine) as s2:
        libs = s2.execute(select(AuditCheckpointLibrary)).scalars().all()
        me = next((l for l in libs if l.industryCode == INDUSTRY_CODE), None)
        print("\n-- persistence check (fresh session) --------------------------")
        if me is None:
            print("  MISSING — the write did NOT land.")
            return 1
        cats = me.categories or []
        n = sum(len(c.get("checkpoints") or []) for c in cats)
        scope = library_subject_scope(me.industryCode, cats)
        print(f"  {me.industryCode}: {n} checkpoint(s), scope={scope}, "
              f"active={me.isActive}, count column={me.checkpointCount}")
        for c in cats:
            print(f"     {c['category_name']:<20} {len(c.get('checkpoints') or []):>3}")

        print("\n-- other libraries unchanged ---------------------------------")
        drift = []
        for l in libs:
            if l.industryCode == INDUSTRY_CODE:
                continue
            was = before.get(l.industryCode)
            same = was == l.checkpointCount
            if not same:
                drift.append(f"{l.industryCode}: {was} -> {l.checkpointCount}")
            print(f"  {l.industryCode:<24} {l.checkpointCount:>5} {'unchanged' if same else 'CHANGED'}")

        ok = n == total and scope == "VENDOR" and me.isActive and not drift
        if drift:
            print(f"\n  DRIFT DETECTED: {drift}")
        print(f"\n  {'VERIFIED' if ok else 'FAILED'}")
        return 0 if ok else 1


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--commit", action="store_true", help="apply (default is a dry run)")
    raise SystemExit(main(ap.parse_args().commit))
