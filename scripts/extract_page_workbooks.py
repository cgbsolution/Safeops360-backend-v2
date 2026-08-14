"""Regenerate the two management-system / social-compliance checkpoint libraries
from the customer's source workbooks.

  • "QMS, EMS OHS.xls"                                  -> page_ims_checkpoints.json
  • "Annexure-2, PIL Social Compliance Audit Checklist" -> page_social_compliance_checkpoints.json

Output matches the shape of app/seed/data/page_industries_checkpoints.json, so
the three libraries are seeded and materialised by exactly the same code path.
Committed alongside the JSON it produces, because the mapping from workbook
geometry to checkpoints (which column is which standard, what "--" means, where
a section banner stops and data starts) is the part worth being able to re-read
when Page issue a new revision of either sheet.

PAGE_IMS is segregated by DEPARTMENT (HR / Admin / OHC), each assessed against
both sheets — see the block above `DEPARTMENTS` for why that is the correct
reading of the workbook and the discipline split was not.

Needs `xlrd` (the QMS workbook is a real BIFF .xls) and `openpyxl`.

Run from the backend root, pointing at the directory holding both workbooks:
    python scripts/extract_page_workbooks.py /path/to/workbooks app/seed/data
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl
import xlrd

SRC = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(".")
OUT = Path(sys.argv[2]) if len(sys.argv) > 2 else Path(__file__).resolve().parents[1] / "app" / "seed" / "data"

IMS_WORKBOOK = "QMS, EMS OHS.xls"
SOCIAL_WORKBOOK = "Annexure-2, PIL Social Compliance Audit Checklist.xlsx"

# ── Statutory classifier ─────────────────────────────────────────────────
# Column I of the Page grading vocabulary is a property of the REQUIREMENT,
# not of the audit, so it is set once here (same reasoning as the existing
# seed_page_industries_library.py). A line naming a licence, consent, NOC,
# statutory register or an evaluation-of-compliance clause is statutory.
STATUTORY = re.compile(
    r"licen[cs]e|consent|\bNOC\b|statutory|legal|compliance obligation|"
    r"evaluation of compliance|certificate|manifest|form \d|register|FSSAI|"
    r"minimum wage|EPFO|ESIC|bonus|child labour|forced labour|discrimination|"
    r"working hours|overtime|standing order|factory licence|PCB|boiler|"
    r"pressure vessel|environmental statement|medical test|health surveillance",
    re.I,
)
# Lines whose failure is a safety / legal exposure rather than a paperwork gap.
CRITICAL = re.compile(
    r"child labour|forced labour|discrimination|fire|emergency|hazard|"
    r"toxic|drowning|electrical safety|evacuation|exit|PPE|personal protective|"
    r"boiler|pressure vessel|HIRA|risk control|safety data sheet|chemical",
    re.I,
)


def classify(text: str) -> tuple[str, str]:
    """-> (criticality, requirement_type)."""
    req = "STATUTORY_REGULATORY" if STATUTORY.search(text) else "INTERNAL_REQUIREMENT"
    if CRITICAL.search(text):
        crit = "critical"
    elif req == "STATUTORY_REGULATORY":
        crit = "major"
    else:
        crit = "minor"
    return crit, req


def clean(v) -> str:
    return re.sub(r"\s+", " ", str(v or "")).strip()


def sl_no(v) -> int | None:
    """The workbook's Sl No, or None when the cell is not one.

    xlrd hands every numeric cell back as a float, so column A reads "1.0" —
    which is why this is a parse and not a `str.isdigit()` check.
    """
    s = clean(v)
    if not re.fullmatch(r"\d+(\.0+)?", s):
        return None
    return int(float(s))


def clean_multiline(v) -> str:
    """Like `clean`, but keeps the line breaks.

    The Audit Reference cells hold an enumerated procedure ("i. … ii. … iii. …")
    one step per line. Collapsing that to a single line turns an auditor's
    method into a paragraph, so the newlines are load-bearing content here.
    """
    lines = [re.sub(r"[ \t]+", " ", ln).strip() for ln in str(v or "").splitlines()]
    return "\n".join(ln for ln in lines if ln)


# ── PAGE_IMS — segregated by DEPARTMENT, not by discipline ───────────────
#
# The first build of this library made the four management-system standards
# (QMS / EMS / OHSMS / EnMS) the segregation axis, because that is how the
# workbook's *columns* are laid out. That was the wrong reading. Page conduct
# ONE audit per DEPARTMENT, and each department is assessed against both sheets:
#
#   Tab 1 "QMS, EMS & OHS"  ->  the IMS stream   (ISO 9001 / 14001 / 45001)
#   Tab 2 "EnMS"            ->  the EnMS stream  (ISO 50001)
#
# Both sheets say so in their own header row 4: "Department : HR, Admin and OHC".
# A discipline-segregated library cannot express that — it produces one QMS
# bucket covering all three departments at once, so an auditor cannot record
# that HR's SOP control is effective while Admin's is not.
#
# So a CATEGORY here is a department, and the three clause columns become
# `standard_clauses` ON the checkpoint instead of three separate copies of it.
# That is the whole restructure; everything below is bookkeeping for it.
#
# `DEPT_`-prefixed rather than bare HR / ADMIN / OHC, and not cosmetically: the
# annual programme resolves a slot's library by which one covers the most
# planned category codes (`programme/materialise._library_for`). PAGE_INDUSTRIES
# already owns a category coded `HR`, so a bare `HR` here would make that
# resolution ambiguous and a slot could silently materialise against the
# internal-audit checklist instead of this one. The code is internal; the
# `category_name` is what any screen shows.
DEPARTMENTS = [
    ("DEPT_HR", "Human Resources", "#7C3AED", "users", "HR"),
    ("DEPT_ADMIN", "Administration", "#0EA5E9", "building-2", "ADMIN"),
    ("DEPT_OHC", "Occupational Health Centre", "#DC2626", "stethoscope", "OHC"),
]

# Clause columns on tab 1 (header row 8). A row is assessed against a standard
# when that standard's clause cell carries a clause; "--" is the workbook's own
# marker for "this line does not apply to this standard".
IMS_CLAUSE_COLUMNS = [
    ("QMS", "ISO 9001:2015", 5),
    ("EMS", "ISO 14001:2015", 6),
    ("OHSMS", "ISO 45001:2018", 7),
]
ENMS_STANDARD = "ISO 50001:2018"

# Tab 1 rows 41–60 sit under the Sewage / Effluent Treatment Plant banners and
# are the Admin department's alone; 1–40 are common to all three. The split is
# the customer's, and it is a NUMBER rather than a keyword match on purpose — a
# banner rename must not silently move twenty checkpoints between departments.
IMS_COMMON_MAX_SL = 40

# Tab 2 carries no plant-specific block, so all 22 EnMS rows are assessed in
# every department.
ENMS_ADMIN_ONLY_SL: frozenset[int] = frozenset()

# Checkpoints that appear on BOTH sheets — the same requirement asked once
# against ISO 9001/14001/45001 and again against ISO 50001. The auditor records
# one finding per stream, so the conduct screen collapses each pair into a
# single card with an IMS / EnMS toggle, and the two reports each take their own
# side of it.
#
# Curated rather than fuzzy-matched. "Master list of documents" (IMS 16) and
# "Master list of documents and formats" (EnMS 12) are the same requirement
# under different wording, which no string comparison gets right; and IMS 19
# splits across EnMS 15+16, which is not a pair at all and must not be forced
# into one. A wrong pair silently hides one of the two findings behind a toggle,
# so this list is the kind of thing that has to be read, not inferred.
IMS_ENMS_PAIRS: list[tuple[int, int]] = [
    (1, 1),    # Previous Audit and NC Closure Status
    (2, 2),    # Departmental Objective (KPI)
    (3, 3),    # Action Plan If KPI Target not achieved
    (4, 4),    # Process Module
    (5, 5),    # Standard Operating Procedure
    (6, 6),    # Needs and expectations of Interested parties, Risk & Opportunities
    (7, 7),    # Organogram (Dept. Organization Chart)
    (15, 11),  # Compliance Obligation / legal and other requirements
    (16, 12),  # Master list of documents (+ formats, on the EnMS sheet)
    (18, 14),  # Continual Improvement
]


def _pair_keys() -> tuple[dict[int, str], dict[int, str]]:
    """Sl No -> pair key, for each stream. The key is shared by the two members
    so the runtime can join them without knowing the mapping."""
    ims: dict[int, str] = {}
    enms: dict[int, str] = {}
    for i, (ims_sl, enms_sl) in enumerate(IMS_ENMS_PAIRS, start=1):
        key = f"PAIR-{i:02d}"
        ims[ims_sl] = key
        enms[enms_sl] = key
    return ims, enms


def _read_ims_rows(sh) -> list[dict]:
    """Tab 1 -> one dict per numbered row, carrying every standard it cites."""
    rows: list[dict] = []
    section = ""
    for r in range(8, sh.nrows):  # row 9 (0-indexed 8) onward
        sl, b = sl_no(sh.cell_value(r, 0)), clean(sh.cell_value(r, 1))
        # Section banners: one of the two cells carries a heading and the other
        # is blank. They scope the lines beneath them (STP / ETP), so they are
        # carried onto the question rather than dropped.
        if not b or sl is None:
            head = b or clean(sh.cell_value(r, 0))
            if head:
                section = head.rstrip(":")
            continue
        clauses = []
        for code, standard, col in IMS_CLAUSE_COLUMNS:
            clause = clean(sh.cell_value(r, col))
            if clause and clause not in {"--", "-", "NA", "N/A"}:
                clauses.append({"code": code, "standard": standard, "clause": clause})
        if not clauses:
            # A numbered row citing no standard at all is not assessable —
            # dropping it silently would be worse than the row never existing,
            # so it is reported by `report()` via the count not matching 60.
            continue
        m = re.match(r"^(Sewage Treatment Plant \(STP\)|Effluent Treatment Plant \(ETP\))", section)
        prefix = f"{'STP' if m and 'STP' in m.group(1) else 'ETP'} — " if m else ""
        rows.append({
            "sl": sl,
            "question": prefix + b,
            "guidance": clean_multiline(sh.cell_value(r, 2)),
            "clauses": clauses,
        })
    return rows


def _read_enms_rows(sh) -> list[dict]:
    """Tab 2 -> one dict per numbered row. Single clause column (F), ISO 50001."""
    rows: list[dict] = []
    for r in range(8, sh.nrows):
        sl, b = sl_no(sh.cell_value(r, 0)), clean(sh.cell_value(r, 1))
        if not b or sl is None:
            continue
        clause = clean(sh.cell_value(r, 5))
        rows.append({
            "sl": sl,
            "question": b,
            "guidance": clean_multiline(sh.cell_value(r, 2)),
            "clauses": (
                [{"code": "EnMS", "standard": ENMS_STANDARD, "clause": clause}]
                if clause else []
            ),
        })
    return rows


def _checkpoint(short: str, stream: str, row: dict, pair_key: str | None) -> dict:
    """One materialisable checkpoint: a workbook row, in a department, on a stream."""
    clauses = row["clauses"]
    crit, req = classify(f"{row['question']} {row['guidance']}")
    return {
        # Coded off the workbook's own Sl No rather than a running counter, so a
        # checkpoint on a report can be traced straight back to a line on the
        # sheet the customer already keeps. Uses the department's SHORT form —
        # `PI-DEPT_HR-IMS-001` reads worse than `PI-HR-IMS-001` and the code is
        # already unique without the disambiguating prefix the category needs.
        "code": f"PI-{short}-{stream}-{row['sl']:03d}",
        "question": row["question"],
        "criticality": crit,
        "requirement_type": req,
        "guidance": row["guidance"],
        # Display strings, for the surfaces that render one line of text.
        "requirement_reference": " · ".join(
            f"{c['code']} {c['clause']}" for c in clauses
        ),
        "standard": " · ".join(c["standard"] for c in clauses),
        # The structured form, which is what the standards rollup aggregates on.
        # An IMS row cites up to three ISO standards at once; collapsing that to
        # the display string above would report "ISO 9001 · ISO 14001 · ISO
        # 45001" as a fourth standard and leave the three real ones empty.
        "standard_clauses": clauses,
        # Which report this checkpoint belongs to. The two reports are separate
        # documents, not two views of one.
        "stream": stream,
        # Same workbook row, in another department — what "replicate this status
        # across departments" copies along.
        "replication_key": f"{stream}-{row['sl']:03d}",
        # Same requirement on the other sheet, in THIS department. Null unless
        # the row is one of IMS_ENMS_PAIRS.
        "pair_key": pair_key,
    }


def build_ims() -> list[dict]:
    wb = xlrd.open_workbook(str(SRC / IMS_WORKBOOK))
    ims_rows = _read_ims_rows(wb.sheet_by_name("QMS, EMS & OHS"))
    enms_rows = _read_enms_rows(wb.sheet_by_name("EnMS"))
    ims_pairs, enms_pairs = _pair_keys()

    cats: list[dict] = []
    for dept, name, colour, icon, short in DEPARTMENTS:
        checkpoints: list[dict] = []
        for row in ims_rows:
            if row["sl"] > IMS_COMMON_MAX_SL and short != "ADMIN":
                continue
            checkpoints.append(_checkpoint(short, "IMS", row, ims_pairs.get(row["sl"])))
        for row in enms_rows:
            if row["sl"] in ENMS_ADMIN_ONLY_SL and short != "ADMIN":
                continue
            checkpoints.append(_checkpoint(short, "ENMS", row, enms_pairs.get(row["sl"])))
        cats.append({
            "category_code": dept,
            "category_name": name,
            "category_color": colour,
            "category_icon": icon,
            # Declared on the category so the runtime classifies this library
            # without a code change, exactly as PAGE_SOCIAL declares its own
            # subject scope and audit category.
            "subject_scope": "OWN_SITE",
            "audit_category": "MANAGEMENT_SYSTEMS",
            # A category here is a DEPARTMENT. The scheduling wizard and the
            # conduct navigator read this to say "Departments in scope" rather
            # than "Disciplines in scope", which is otherwise a lie on screen.
            "segregation": "DEPARTMENT",
            # Conformance / Non-Conformance / Observation — the three parameters
            # the customer's sheet actually carries, in place of the engine's
            # seven-value status ladder. See services/page_grading.TRISTATE.
            "conformance_mode": "TRISTATE",
            "streams": ["IMS", "ENMS"],
            "checkpoints": checkpoints,
        })
    return cats


# ── PAGE_SOCIAL ──────────────────────────────────────────────────────────
# Column C is the section (= discipline), column E the checkpoint. Both are
# vertically merged, so the section is carried down until it changes.
#
# Every section of this checklist is backed by statute — the Factories Act,
# Payment of Wages / Minimum Wages / Payment of Bonus Acts, the EPF & ESI
# Acts, the Child & Adolescent Labour (P&R) Act, the Trade Unions Act and the
# Environment (Protection) Act respectively. `requirement_type` is therefore
# STATUTORY_REGULATORY throughout, which is a statement about the checklist,
# not a shortcut: there is no internal-only line in it.
#
# The per-section base criticality carries the part that DOES differ — the
# zero-tolerance sections (child / forced labour, discrimination) and life
# safety are critical whatever the wording of the individual line, and the
# CRITICAL keyword pass can lift a line but never lower one.
SOCIAL_STYLE = {
    "LAWS": ("#7C3AED", "scale", "major"),
    "HOURS": ("#0EA5E9", "clock", "major"),
    "WAGES": ("#16A34A", "wallet", "major"),
    "HS": ("#DC2626", "shield", "critical"),
    "FOA": ("#EA580C", "users", "major"),
    "CHILD": ("#BE123C", "user-x", "critical"),
    "DISCRIM": ("#DB2777", "user-minus", "critical"),
    "FORCED": ("#9333EA", "lock", "critical"),
    "ENV": ("#059669", "leaf", "major"),
}
SOCIAL_CODES = ["LAWS", "HOURS", "WAGES", "HS", "FOA", "CHILD", "DISCRIM", "FORCED", "ENV"]


def build_social() -> list[dict]:
    wb = openpyxl.load_workbook(SRC / SOCIAL_WORKBOOK, data_only=True)
    sh = wb["Sheet1"]
    cats: list[dict] = []
    section = None
    for r in range(5, sh.max_row + 1):
        sec = clean(sh.cell(r, 3).value)
        q = clean(sh.cell(r, 5).value)
        if sec and sec != section:
            section = sec
            code = SOCIAL_CODES[len(cats)]
            colour, icon, _base = SOCIAL_STYLE[code]
            cats.append({
                "category_code": code, "category_name": sec,
                "category_color": colour, "category_icon": icon,
                # Declared on every category so `library_subject_scope` and
                # `library_audit_category` both classify this library without a
                # code change. This checklist audits a SUPPLIER's factory —
                # licences, wages, child labour — so it belongs to the vendor
                # subject, not to an audit of our own site.
                "subject_scope": "VENDOR",
                "audit_category": "SOCIAL_COMPLIANCE",
                "checkpoints": [],
            })
        if not q or not cats:
            continue
        cat = cats[-1]
        base = SOCIAL_STYLE[cat["category_code"]][2]
        crit = "critical" if base == "critical" or CRITICAL.search(q) else base
        cat["checkpoints"].append({
            "code": f"PI-SC-{cat['category_code']}-{len(cat['checkpoints']) + 1:03d}",
            "question": q,
            "criticality": crit,
            "requirement_type": "STATUTORY_REGULATORY",
            "guidance": "",
            "requirement_reference": cat["category_name"],
            "standard": "PIL Social Compliance Audit Checklist (Annexure-2, v4)",
        })
    return cats


def report(label: str, cats: list[dict]) -> None:
    total = sum(len(c["checkpoints"]) for c in cats)
    axis = "departments" if cats and cats[0].get("segregation") == "DEPARTMENT" else "disciplines"
    print(f"\n{label} — {total} checkpoints / {len(cats)} {axis}")
    for c in cats:
        cps = c["checkpoints"]
        stat = sum(1 for cp in cps if cp["requirement_type"] == "STATUTORY_REGULATORY")
        streams = ""
        if any(cp.get("stream") for cp in cps):
            ims = sum(1 for cp in cps if cp.get("stream") == "IMS")
            enms = sum(1 for cp in cps if cp.get("stream") == "ENMS")
            pairs = len({cp["pair_key"] for cp in cps if cp.get("pair_key")})
            streams = f"  [IMS {ims} + EnMS {enms}, {pairs} paired]"
        print(f"  {c['category_code']:<8} {c['category_name'][:40]:<42} "
              f"{len(cps):>3}  ({stat} statutory){streams}")


if __name__ == "__main__":
    ims, social = build_ims(), build_social()
    report("PAGE_IMS", ims)
    report("PAGE_SOCIAL", social)
    OUT.mkdir(parents=True, exist_ok=True)
    (OUT / "page_ims_checkpoints.json").write_text(
        json.dumps(ims, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    (OUT / "page_social_compliance_checkpoints.json").write_text(
        json.dumps(social, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"\nWrote to {OUT.resolve()}")
