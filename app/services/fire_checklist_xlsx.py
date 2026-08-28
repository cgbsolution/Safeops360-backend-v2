"""Excel (.xlsx) export for the Page Industries fire checklists and register.

The sibling of `fire_checklist_pdf.py`, and deliberately a different artefact
rather than the same one in another wrapper:

  PDF    the controlled document. Handed to an external auditor, so fidelity to
         the paper original beats everything — fixed columns, the sign-off block
         drawn where the sheet prints it, nothing reflowable.
  XLSX   the working copy. Handed to an engineer who is going to sort the
         register by refill due date, filter the overdue cylinders, or paste a
         month of ticks into a report. It carries the same rows and the same
         controlled-document header, but as data: real dates, one row per record,
         freeze panes and an autofilter.

Both read the SAME payloads the API already builds (`grid_out`, `run_out`,
`build_register`), so an export can never disagree with the screen that produced
it — the two renderers are two views of one dict, not two queries.

Three renderers because the workbooks have three shapes, matching the PDF module
one for one:
  render_grid     — items x periods (daily month page, FE year page, quarter page)
  render_form     — one period, sectioned (monthly / annual sheets)
  render_register — PIL/EHSD/CL/028-R1, the sixteen columns

WHY DATES ARE WRITTEN AS DATES
------------------------------
The register's whole purpose is "what is overdue", and a column of dd.mm.yyyy
*strings* sorts 01.02.2031 above 27.04.2026. Real date cells with a display
format are what make the file useful the moment it opens. The one exception is a
date that was never recorded: that cell keeps the register's own words, "not
recorded", because a blank would read as "nothing due" and a register that paints
its own gaps as compliance is worse than no register at all.
"""

from __future__ import annotations

import re
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.services.report_pdf import REPORT_TZ, REPORT_TZ_LABEL

# Midnight Executive, as ARGB — the same inks as fire_checklist_pdf.py and
# src/app/(dashboard)/fire-safety/lib.ts, so the workbook, the PDF and the screen
# read as one system. A literal hex anywhere below is a bug for the same reason
# it is in the other two.
NAVY = "FF0B1F4D"
GOLD = "FFC9A961"
GOLD_SOFT = "FFE8D9B0"
ICE = "FFE8EEF7"
INK = "FF1A202C"
GREY = "FF5A6273"
WHITE = "FFFFFFFF"
RED = "FFC0392B"
AMBER = "FFC88214"
GREEN = "FF2E7D5B"

# Answer inks, matching the PDF's CELL_TINT / CELL_INK pair.
ANSWER_FILL = {"YES": "FFEEF6F1", "NO": "FFFCEBE9", "NA": "FFF0F2F6"}
ANSWER_INK = {"YES": GREEN, "NO": RED, "NA": GREY}

BADGE_INK = {"OVERDUE": RED, "DUE_SOON": AMBER, "OK": GREEN, "NOT_RECORDED": GREY}

DATE_FORMAT = "DD.MM.YYYY"  # the format the source sheets print

_THIN = Side(style="thin", color="FFD3DEEE")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# openpyxl refuses control characters outright; a stray one in transcribed sheet
# text would fail the whole export rather than the one cell.
_CONTROL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _clean(v: Any) -> str:
    return _CONTROL.sub("", "" if v is None else str(v))


def _sheet_name(name: str) -> str:
    """Excel forbids : \\ / ? * [ ] in a tab name and caps it at 31 chars."""
    return re.sub(r"[:\\/?*\[\]]", "-", _clean(name))[:31] or "Sheet1"


def _as_date(iso: str | None) -> date | None:
    if not iso:
        return None
    try:
        return datetime.fromisoformat(str(iso).replace("Z", "+00:00")).date()
    except ValueError:
        return None


def _now_label() -> str:
    return f"{datetime.now(REPORT_TZ).strftime('%d-%m-%Y %H:%M')} {REPORT_TZ_LABEL}"


# ── the controlled-document header block ─────────────────────────────────────
def _title_block(ws: Worksheet, doc: dict[str, Any], title: str, width: int, subtitle: str = "") -> int:
    """The same five fields the PDF prints across the top of every sheet, and the
    reason an auditor can put the two side by side: document number, what it
    supersedes, effective date, review date, revision. Returns the next free row.
    """
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    c = ws.cell(row=1, column=1, value=_clean(title))
    c.font = Font(bold=True, size=13, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    org = ws.cell(row=2, column=1, value=_clean(f"PAGE INDUSTRIES LIMITED · {doc.get('department', 'EHS')}"))
    org.font = Font(bold=True, size=9, color=GOLD)
    org.fill = PatternFill("solid", fgColor=NAVY)
    org.alignment = Alignment(horizontal="center")

    fields = [
        ("Document No.", doc.get("documentNo")),
        ("Supersedes No.", doc.get("supersedesNo")),
        ("Effective Date", doc.get("effectiveDate")),
        ("Review Date", doc.get("reviewDate")),
        ("Revision", doc.get("revision")),
    ]
    row = 3
    for i, (label, value) in enumerate(fields):
        lc = ws.cell(row=row, column=i * 2 + 1, value=label)
        lc.font = Font(bold=True, size=8, color=NAVY)
        lc.fill = PatternFill("solid", fgColor=ICE)
        lc.border = BORDER
        d = _as_date(value) if "Date" in label else None
        vc = ws.cell(row=row, column=i * 2 + 2, value=d or _clean(value) or "—")
        if d:
            vc.number_format = DATE_FORMAT
        vc.font = Font(size=8, color=INK)
        vc.border = BORDER
    row += 1

    if subtitle:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)
        s = ws.cell(row=row, column=1, value=_clean(subtitle))
        s.font = Font(size=8.5, color=GREY)
        row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)
    g = ws.cell(row=row, column=1, value=f"Generated {_now_label()} · SafeOps360")
    g.font = Font(size=7.5, italic=True, color=GREY)
    return row + 2


def _header_row(ws: Worksheet, row: int, labels: list[str], *, fills: list[str] | None = None) -> None:
    for i, label in enumerate(labels, start=1):
        c = ws.cell(row=row, column=i, value=_clean(label))
        c.font = Font(bold=True, size=8, color=NAVY)
        c.fill = PatternFill("solid", fgColor=(fills[i - 1] if fills else ICE))
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = 26


def _widths(ws: Worksheet, widths: list[float]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _sign_off_rows(ws: Worksheet, row: int, doc: dict[str, Any], sign_off: dict[str, Any] | None) -> int:
    """The three-stage block the sheets print. Reproduced because it is what the
    auditor checks: who prepared, who reviewed, who approved, and when.
    """
    roles = doc.get("signOffRoles") or [
        "Prepared by: Person In-charge",
        "Reviewed by: Intermediatory Head",
        "Approved by: HOD",
    ]
    so = sign_off or {}
    stamps = [
        (so.get("preparedByName"), so.get("preparedAt")),
        (so.get("reviewedByName"), so.get("reviewedAt")),
        (so.get("approvedByName"), so.get("approvedAt")),
    ]
    row += 1
    for i, role in enumerate(roles[:3]):
        name, at = stamps[i] if i < len(stamps) else (None, None)
        rc = ws.cell(row=row, column=1, value=_clean(role))
        rc.font = Font(bold=True, size=8, color=NAVY)
        rc.fill = PatternFill("solid", fgColor=ICE)
        rc.border = BORDER
        when = _as_date(at)
        vc = ws.cell(
            row=row,
            column=2,
            value=_clean(f"{name} · {when.strftime('%d.%m.%Y')}") if name and when
            else _clean(name) if name
            else "Sign. & Date:",
        )
        vc.font = Font(size=8, color=INK if name else GREY)
        vc.border = BORDER
        row += 1
    return row


def _footnotes(ws: Worksheet, row: int, lines: list[str] | None, width: int) -> int:
    for line in lines or []:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(width, 2))
        c = ws.cell(row=row, column=1, value=_clean(line))
        c.font = Font(size=7.5, italic=True, color=GREY)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    return row


def _save(wb: Workbook) -> bytes:
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ═══════════════════════════════════════════════════════════════════════════
# render_grid — items x periods
# ═══════════════════════════════════════════════════════════════════════════
def render_grid(payload: dict[str, Any]) -> bytes:
    """The daily month page, the FE year page and the quarter page, as a sheet.

    One row per check, one column per period — the same shape the screen and the
    PDF use, so a cell an inspector can point to on the wall chart is the same
    cell here.
    """
    doc = payload.get("document") or {}
    cols = payload.get("columns") or []
    rows = payload.get("rows") or []

    wb = Workbook()
    ws = wb.active
    ws.title = _sheet_name(payload.get("window") or "Grid")

    bits = [f"Asset: {payload.get('assetCode', '')}"]
    if payload.get("allottedSerialNo"):
        bits.append(f"FE No.: {payload['allottedSerialNo']}")
    if payload.get("assetSubtype"):
        bits.append(f"Type: {payload['assetSubtype']}")
    if payload.get("assetLocation"):
        bits.append(f"Location & Floor: {payload['assetLocation']}")
    bits.append(f"Period: {payload.get('window', '')}")

    width = len(cols) + 2
    row = _title_block(ws, doc, payload.get("templateName", "Fire checklist"), width, "   |   ".join(bits))

    head_row = row
    # A non-working day is tinted at the header, exactly as the PDF does it, so a
    # whole empty column reads as "the plant was shut" and not "31 missed checks".
    fills = [ICE, ICE] + [GOLD_SOFT if c.get("nonWorkingDay") else ICE for c in cols]
    _header_row(ws, head_row, ["Sl.", "Checks to be done"] + [str(c.get("header", "")) for c in cols], fills=fills)
    row += 1

    for n, r in enumerate(rows, start=1):
        text = _clean(r.get("text"))
        if r.get("guidance"):
            text = f"{text}  (Note: {r['guidance']})"
        sl = ws.cell(row=row, column=1, value=n)
        sl.font = Font(size=8, color=GREY)
        sl.alignment = Alignment(horizontal="center")
        sl.border = BORDER
        tc = ws.cell(row=row, column=2, value=text)
        tc.font = Font(size=8, color=INK)
        tc.alignment = Alignment(wrap_text=True, vertical="top")
        tc.border = BORDER

        cells = r.get("cells") or {}
        for j, col in enumerate(cols, start=3):
            cell_data = cells.get(col.get("periodLabel")) or {}
            value = cell_data.get("value")
            key = str(value).upper() if value else None
            c = ws.cell(row=row, column=j, value=_clean(value) if value else "")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = BORDER
            # The remark, attached to the cell it belongs to. Excel shows it on
            # hover and flags the cell, so the grid keeps its one-glance shape and
            # the comment is still one hover away rather than only in the block
            # at the foot of the sheet.
            note = (cell_data.get("note") or "").strip()
            if note:
                c.comment = Comment(note, "SafeOps360")
            if key in ANSWER_FILL:
                c.fill = PatternFill("solid", fgColor=ANSWER_FILL[key])
                c.font = Font(bold=key == "NO", size=8, color=ANSWER_INK[key])
            else:
                c.font = Font(size=8, color=GREY)
                if col.get("nonWorkingDay"):
                    c.fill = PatternFill("solid", fgColor=GOLD_SOFT)
        row += 1

    # The per-period stage strip. A grid page covers many runs, each with its own
    # sign-off state, so one foot block cannot speak for all of them — this row is
    # what answers "which days are approved?".
    sc = ws.cell(row=row, column=2, value="Sign-off stage")
    sc.font = Font(bold=True, size=8, color=NAVY)
    sc.fill = PatternFill("solid", fgColor=ICE)
    sc.border = BORDER
    for j, col in enumerate(cols, start=3):
        c = ws.cell(row=row, column=j, value=_clean(col.get("stage")) or "—")
        c.font = Font(size=7.5, color=NAVY if col.get("stage") else GREY)
        c.alignment = Alignment(horizontal="center")
        c.fill = PatternFill("solid", fgColor=ICE)
        c.border = BORDER
    row += 2

    # "Comments on the back side of this page" — the sheet's own footnote. A grid
    # cell cannot hold the remark that explains a "No", and on the paper original
    # it does not; this is the back of the page. A cell that carries one is also
    # marked in place, so the two are findable from each other.
    remarks: list[tuple[str, str, str]] = []
    header_by_period = {c.get("periodLabel"): str(c.get("header", c.get("periodLabel", ""))) for c in cols}
    for n, r in enumerate(rows, start=1):
        for period, cell in (r.get("cells") or {}).items():
            note = (cell or {}).get("note")
            if note and str(note).strip():
                remarks.append((header_by_period.get(period, period), f"{n}. {r.get('text', '')}", str(note).strip()))

    if remarks:
        rh = ws.cell(row=row, column=1, value="Remarks")
        rh.font = Font(bold=True, size=9, color=NAVY)
        rh.fill = PatternFill("solid", fgColor=ICE)
        row += 1
        _header_row(ws, row, ["Period", "Check", "Remark"] + [""] * max(0, width - 3))
        row += 1
        for period, item, note in remarks:
            for i, v in enumerate((period, item, note), start=1):
                c = ws.cell(row=row, column=i, value=_clean(v))
                c.font = Font(size=8, color=NAVY if i == 1 else INK)
                c.alignment = Alignment(wrap_text=i > 1, vertical="top")
                c.border = BORDER
            row += 1
        row += 1

    row = _footnotes(ws, row, doc.get("footnotes"), width)
    _sign_off_rows(ws, row, doc, None)

    _widths(ws, [5, 52] + [7.5 if len(cols) > 12 else 11] * len(cols))
    # Freeze below the header and right of the wording: scrolling to 31 December
    # must not scroll the question off the screen.
    ws.freeze_panes = ws.cell(row=head_row + 1, column=3)
    return _save(wb)


# ═══════════════════════════════════════════════════════════════════════════
# render_form — one period, sectioned
# ═══════════════════════════════════════════════════════════════════════════
def render_form(payload: dict[str, Any]) -> bytes:
    """The monthly / annual sheets: headings, items, answers, notes, sign-off."""
    doc = payload.get("document") or {}

    wb = Workbook()
    ws = wb.active
    ws.title = _sheet_name(payload.get("periodLabel") or "Checklist")

    bits = [f"Asset: {payload.get('assetCode', '')}"]
    if payload.get("assetLocation"):
        bits.append(f"Location: {payload['assetLocation']}")
    bits.append(f"Period: {payload.get('periodLabel', '')}")
    bits.append(f"Status: {payload.get('stage', '')}")

    width = 4
    row = _title_block(ws, doc, payload.get("templateName", "Fire checklist"), width, "   |   ".join(bits))

    head_row = row
    _header_row(ws, head_row, ["Sl.", "Checks to be done", "Observation", "Remarks"])
    row += 1

    for sec in payload.get("sections") or []:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)
        t = ws.cell(row=row, column=1, value=_clean(sec.get("title")))
        t.font = Font(bold=True, size=8.5, color=NAVY)
        t.fill = PatternFill("solid", fgColor=ICE)
        t.border = BORDER
        row += 1
        if sec.get("note"):
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)
            n = ws.cell(row=row, column=1, value=_clean(sec["note"]))
            n.font = Font(size=7.5, italic=True, color=GREY)
            row += 1

        for i, item in enumerate(sec.get("items") or [], start=1):
            text = _clean(item.get("text"))
            if item.get("guidance"):
                text = f"{text}  (Note: {item['guidance']})"
            sl = ws.cell(row=row, column=1, value=i)
            sl.font = Font(size=8, color=GREY)
            sl.alignment = Alignment(horizontal="center")
            sl.border = BORDER
            tc = ws.cell(row=row, column=2, value=text)
            tc.font = Font(size=8, color=INK)
            tc.alignment = Alignment(wrap_text=True, vertical="top")
            tc.border = BORDER

            value = item.get("value")
            key = str(value).upper() if value else None
            vc = ws.cell(row=row, column=3, value=_clean(value) if value else "")
            vc.alignment = Alignment(horizontal="center", vertical="center")
            vc.border = BORDER
            if key in ANSWER_FILL:
                vc.fill = PatternFill("solid", fgColor=ANSWER_FILL[key])
                vc.font = Font(bold=key == "NO", size=8, color=ANSWER_INK[key])
            else:
                vc.font = Font(size=8, color=GREY)

            nc = ws.cell(row=row, column=4, value=_clean(item.get("note")))
            nc.font = Font(size=8, color=GREY)
            nc.alignment = Alignment(wrap_text=True, vertical="top")
            nc.border = BORDER
            row += 1

    row += 1
    row = _footnotes(ws, row, doc.get("footnotes"), width)
    _sign_off_rows(ws, row, doc, payload.get("signOff"))

    _widths(ws, [5, 72, 14, 34])
    ws.freeze_panes = ws.cell(row=head_row + 1, column=1)
    return _save(wb)


# ═══════════════════════════════════════════════════════════════════════════
# render_register — PIL/EHSD/CL/028-R1
# ═══════════════════════════════════════════════════════════════════════════
# Key, label, column width. Same sixteen columns in the same order as the source
# sheet and the PDF — an auditor reading one and an engineer reading the other
# are reading the same register.
_REG_COLS: list[tuple[str, str, float]] = [
    ("slNo", "Sl.", 5), ("serialNo", "Mfr Serial No.", 16), ("type", "Type", 11),
    ("capacity", "Capacity", 11), ("yearOfManufacture", "Yr Mfg", 8),
    ("expiryDate", "Expiry Date", 13), ("make", "Make", 14),
    ("allottedSerialNo", "Alloted Serial No.", 16), ("location", "Location", 30),
    ("hpTestedOn", "HP tested on", 13), ("hpTestDueDate", "HP Test due", 13),
    ("dateOfDischarge", "Discharged", 13), ("refilledOn", "Refilled on", 13),
    ("dueForRefilling", "Due for refilling", 15), ("weightKg", "Wt (kg)", 8),
    ("remarks", "Remarks", 28),
]

_DATE_KEYS = {"expiryDate", "hpTestedOn", "hpTestDueDate", "dateOfDischarge",
              "refilledOn", "dueForRefilling"}

# Which badge governs which column, so the colour sits on the date it is about
# rather than on the row — the same mapping the PDF uses, and what makes "which
# of the three is overdue?" answerable from the file.
_COL_BADGE = {"expiryDate": "cylinderLife", "hpTestDueDate": "hpTest", "dueForRefilling": "refill"}


def render_register(payload: dict[str, Any]) -> bytes:
    doc = payload.get("document") or {}
    rows = payload.get("rows") or []
    summary = payload.get("summary") or {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Register"

    subtitle = (
        f"{summary.get('total', len(rows))} cylinder(s)   |   "
        f"Overdue: {summary.get('overdue', 0)}   |   Due within 30 days: {summary.get('dueSoon', 0)}   |   "
        f"Date not recorded: {summary.get('notRecorded', 0)}"
    )
    width = len(_REG_COLS)
    row = _title_block(ws, doc, doc.get("title") or "REGISTER OF FIRE EXTINGUISHERS", width, subtitle)

    head_row = row
    _header_row(ws, head_row, [label for _k, label, _w in _REG_COLS])
    row += 1

    for r in rows:
        badges = r.get("badges") or {}
        for i, (key, _label, _w) in enumerate(_REG_COLS, start=1):
            raw = r.get(key)
            badge_key = _COL_BADGE.get(key)
            status = (badges.get(badge_key) or {}).get("status") if badge_key else None

            if key in _DATE_KEYS:
                d = _as_date(raw)
                # "not recorded" rather than a blank: a cylinder with no refill
                # date on file is a gap in the register, and an empty cell reads
                # as "nothing due".
                value = d if d else ("not recorded" if status == "NOT_RECORDED" else "")
            else:
                value = raw if isinstance(raw, (int, float)) else _clean(raw)

            c = ws.cell(row=row, column=i, value=value)
            c.border = BORDER
            if isinstance(value, date):
                c.number_format = DATE_FORMAT
            if status:
                c.font = Font(bold=status in ("OVERDUE", "DUE_SOON"), size=8,
                              color=BADGE_INK.get(status, INK))
            else:
                c.font = Font(size=8, color=INK)
            c.alignment = Alignment(
                horizontal="left" if key in ("location", "remarks") else "center",
                vertical="center",
                wrap_text=key in ("location", "remarks"),
            )
        row += 1

    # The autofilter is the reason this file exists rather than a second PDF:
    # "show me everything overdue on refill" is one click, not a re-run.
    if rows:
        ws.auto_filter.ref = (
            f"A{head_row}:{get_column_letter(width)}{row - 1}"
        )

    row += 1
    row = _footnotes(ws, row, [
        "Badge key: red (past due) / bold amber (due within 30 days) / green (in date) / "
        "'not recorded' (no date on file — a register gap, not compliance).",
        "HP test and refill dates are held as asset certificates; this sheet projects the current "
        "certificate of each type.",
    ], width)
    _sign_off_rows(ws, row, doc, None)

    _widths(ws, [w for _k, _l, w in _REG_COLS])
    ws.freeze_panes = ws.cell(row=head_row + 1, column=1)
    return _save(wb)


# ═══════════════════════════════════════════════════════════════════════════
# render_assets — the "All other fire assets" tab
# ═══════════════════════════════════════════════════════════════════════════
_ASSET_COLS: list[tuple[str, str, float]] = [
    ("equipmentCode", "Code", 18), ("type", "Type", 20), ("assetSubtype", "Subtype", 14),
    ("location", "Location", 32), ("capacitySpec", "Capacity / spec", 16),
    ("make", "Make", 14), ("model", "Model", 14), ("serialNo", "Serial no.", 16),
    ("maintenanceContractor", "Maintenance contractor", 24),
    ("lastInspectionDate", "Last inspected", 14),
    ("nextInspectionDueDate", "Next due", 14), ("status", "Status", 16),
]

_STATUS_INK = {
    "ACTIVE": GREEN, "DUE_INSPECTION": AMBER, "OVERDUE": RED,
    "NON_COMPLIANT": RED, "OUT_OF_SERVICE": GREY, "DECOMMISSIONED": GREY,
}


def render_assets(rows: list[dict[str, Any]], *, title: str = "FIRE ASSET REGISTER") -> bytes:
    """Panels, hydrants, hose reels, detectors, emergency lights.

    The asset types the controlled sixteen-column sheet does not cover, which had
    no export at all — the one tab of the register an engineer could not take off
    the screen.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Fire assets"

    counts: dict[str, int] = {}
    for r in rows:
        s = str(r.get("status") or "")
        counts[s] = counts.get(s, 0) + 1
    subtitle = f"{len(rows)} asset(s)   |   " + "   |   ".join(
        f"{k.replace('_', ' ').title()}: {v}" for k, v in sorted(counts.items())
    )

    width = len(_ASSET_COLS)
    row = _title_block(ws, {"department": "EHS"}, title, width, subtitle)

    head_row = row
    _header_row(ws, head_row, [label for _k, label, _w in _ASSET_COLS])
    row += 1

    for r in rows:
        for i, (key, _label, _w) in enumerate(_ASSET_COLS, start=1):
            raw = r.get(key)
            if key in ("lastInspectionDate", "nextInspectionDueDate"):
                value = _as_date(raw) or ""
            elif key in ("type", "status"):
                value = _clean(raw).replace("_", " ")
            else:
                value = _clean(raw)
            c = ws.cell(row=row, column=i, value=value)
            c.border = BORDER
            if isinstance(value, date):
                c.number_format = DATE_FORMAT
            ink = _STATUS_INK.get(str(r.get("status") or ""), INK) if key == "status" else INK
            c.font = Font(bold=key == "status", size=8, color=ink)
            c.alignment = Alignment(
                horizontal="left" if key in ("location", "maintenanceContractor") else "center",
                vertical="center",
                wrap_text=key in ("location", "maintenanceContractor"),
            )
        row += 1

    if rows:
        ws.auto_filter.ref = f"A{head_row}:{get_column_letter(width)}{row - 1}"

    row += 1
    _footnotes(ws, row, [
        "Status is computed nightly from each asset's inspection due date. Overrides, "
        "out-of-service and frequency changes live on the asset detail page.",
    ], width)

    _widths(ws, [w for _k, _l, w in _ASSET_COLS])
    ws.freeze_panes = ws.cell(row=head_row + 1, column=1)
    return _save(wb)


__all__ = ["render_grid", "render_form", "render_register", "render_assets", "MEDIA_TYPE"]


# ═══════════════════════════════════════════════════════════════════════════
# Generic register (pdfTemplateKey = GENERIC_REGISTER)
# ═══════════════════════════════════════════════════════════════════════════
def render_generic_register(payload: dict[str, Any]) -> bytes:
    """Any config-driven register as a workbook, with the autofilter.

    `render_register` above stays as it is — its columns are transcribed from
    the client's controlled sheet and must not move. This renders whatever
    `document.columns` says, which is what makes a new register a seed row
    rather than a second exporter.

    The autofilter is the point of having this next to the PDF at all: "show me
    every panel overdue" is one click here and a re-run of the export there.
    """
    doc = payload.get("document") or {}
    rows = payload.get("rows") or []
    summary = payload.get("summary") or {}
    columns = [tuple(c) for c in (doc.get("columns") or []) if c]
    if not columns:
        columns = [("equipmentCode", "Code"), ("location", "Location"), ("status", "Status")]

    wb = Workbook()
    ws = wb.active
    ws.title = _sheet_name("Register")

    subtitle = (
        f"{summary.get('total', len(rows))} asset(s)   |   "
        f"Overdue: {summary.get('overdue', 0)}   |   Due within 30 days: {summary.get('dueSoon', 0)}   |   "
        f"Date not recorded: {summary.get('notRecorded', 0)}"
    )
    row = _title_block(ws, doc, doc.get("title") or "FIRE ASSET REGISTER", len(columns), subtitle)

    head_row = row
    for i, (_key, label) in enumerate(columns, start=1):
        c = ws.cell(row=head_row, column=i, value=_clean(label))
        c.font = Font(bold=True, size=9, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r_i, data in enumerate(rows, start=head_row + 1):
        for c_i, (key, _label) in enumerate(columns, start=1):
            value = data.get(key)
            # Dates land as real dates, not strings — otherwise sorting the
            # register by due date sorts it lexically, which is the one thing
            # someone opens this file to do.
            as_date = _as_date(value) if isinstance(value, str) else None
            cell = ws.cell(row=r_i, column=c_i, value=as_date if as_date else value)
            if as_date:
                cell.number_format = "DD.MM.YYYY"
            cell.alignment = Alignment(vertical="top", wrap_text=key in ("location", "remarks"))

    last_row = head_row + len(rows)
    ws.auto_filter.ref = f"A{head_row}:{ws.cell(row=head_row, column=len(columns)).coordinate[0:1]}{max(last_row, head_row)}"
    ws.freeze_panes = ws.cell(row=head_row + 1, column=1)
    for i, (key, label) in enumerate(columns, start=1):
        width = 34 if key in ("location", "remarks") else max(12, min(24, len(str(label)) + 6))
        ws.column_dimensions[ws.cell(row=head_row, column=i).column_letter].width = width

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
